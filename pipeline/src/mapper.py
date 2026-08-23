"""Auto-map the files in a workspace's raw/ folder to pipeline roles.

Produces / refreshes the workspace manifest.yaml:
    company_docs:   {company_briefing: {file, format}, ...}
    markets:        {Spain: {screening_report: {...}, facility_register: {...}}, ...}
    shared_docs:    {master_registry: {...}}
    unassigned:     [files present in raw/ that matched no role]

Rules are filename heuristics + column-header sniffing for facility registers.
A human can edit manifest.yaml at any time; re-mapping only fills blanks and
refreshes `unassigned` - it never overwrites an explicit assignment.
"""
import csv
import re
from pathlib import Path

import openpyxl
import yaml

FORMATS = {".pdf": "pdf", ".xlsx": "xlsx", ".html": "html", ".htm": "html", ".csv": "csv"}

COUNTRY_TOKENS = {
    "portugal": "Portugal", "germany": "Germany", "netherlands": "Netherlands",
    "poland": "Poland", "spain": "Spain", "france": "France", "italy": "Italy",
    "austria": "Austria", "belgium": "Belgium", "sweden": "Sweden", "denmark": "Denmark",
    "norway": "Norway", "finland": "Finland", "ireland": "Ireland", "greece": "Greece",
    "czech": "Czechia", "switzerland": "Switzerland", "uk": "United Kingdom",
}

COLUMN_SYNONYMS = {
    "id": ["facility_id", "site_id", "unit_ref", "registry_ref", "id", "ref", "code"],
    "name": ["facility_name", "site_name", "canonical_name", "name", "unit_name"],
    "country": ["country", "nation", "land"],
    "type": ["type", "category", "unit_type", "kind"],
    "capacity": ["annual_colonoscopy_capacity", "capacity_annual", "capacity_per_year",
                 "capacity", "annual_capacity", "volume"],
    "period": ["reporting_period", "period", "frequency"],
    "note": ["note", "notes", "comment"],
}


def _detect_columns(headers: list[str]) -> dict:
    cols = {}
    lower = {h.lower().strip(): h for h in headers if h}
    for role, syns in COLUMN_SYNONYMS.items():
        for s in syns:
            if s in lower:
                cols[role] = lower[s]
                break
    return cols


def _headers(path: Path, fmt: str) -> list[str]:
    try:
        if fmt == "csv":
            with open(path) as f:
                return next(csv.reader(f))
        if fmt == "xlsx":
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb.active
            return [str(c.value) for c in next(ws.iter_rows(max_row=1))]
        if fmt == "html":
            import re as _re
            head = open(path).read()
            ths = _re.findall(r"<th[^>]*>(.*?)</th>", head, _re.S)
            return [_re.sub(r"<[^>]+>", "", t).strip() for t in ths[:10]]
    except Exception:  # noqa: BLE001
        pass
    return []


def _country_of(fname: str) -> str | None:
    low = fname.lower()
    for tok, country in COUNTRY_TOKENS.items():
        if tok in low:
            return country
    return None


def classify(path: Path) -> tuple[str, dict | None]:
    """-> (role, spec) where role in company_docs/markets/shared/skip/unassigned."""
    fname = path.name
    low = fname.lower()
    fmt = FORMATS.get(path.suffix.lower())
    if fmt is None:
        return "unassigned", None
    spec = {"file": fname, "format": fmt}
    if "candidate" in low and "brief" in low:
        return "skip", None                      # the exercise brief, not source data
    if "company" in low and "brief" in low:
        return "company_docs:company_briefing", spec
    if "landscape" in low:
        return "company_docs:market_landscape", spec
    if "funding" in low:
        return "company_docs:funding_call", spec
    if "competitor" in low or "oncostream" in low:
        return "company_docs:competitor_brief", spec
    if "master" in low and ("registr" in low or "list" in low):
        cols = _detect_columns(_headers(path, fmt))
        if cols:
            spec["columns"] = cols
        return "shared_docs:master_registry", spec
    country = _country_of(low)
    if "screening" in low and country:
        return f"markets:{country}:screening_report", spec
    if ("facilit" in low or "register" in low or "units" in low) and country:
        cols = _detect_columns(_headers(path, fmt))
        if cols:
            spec["columns"] = cols
        if fmt == "xlsx":
            try:
                wb = openpyxl.load_workbook(path, read_only=True)
                spec["sheet"] = wb.sheetnames[0]
            except Exception:  # noqa: BLE001
                pass
        return f"markets:{country}:facility_register", spec
    return "unassigned", None


def build_manifest(ws: Path) -> dict:
    """Refresh <ws>/manifest.yaml from raw/ contents. Explicit entries win."""
    mpath = ws / "manifest.yaml"
    manifest = {"company_docs": {}, "markets": {}, "shared_docs": {}, "skipped": [], "unassigned": []}
    if mpath.exists():
        existing = yaml.safe_load(open(mpath)) or {}
        for k in ("company_docs", "markets", "shared_docs"):
            manifest[k] = existing.get(k) or {}

    assigned_files = set()
    for section in ("company_docs", "shared_docs"):
        assigned_files |= {v["file"] for v in manifest[section].values() if isinstance(v, dict)}
    for m in manifest["markets"].values():
        assigned_files |= {v["file"] for v in m.values() if isinstance(v, dict)}

    for path in sorted((ws / "raw").iterdir()):
        if not path.is_file() or path.name.startswith("."):
            continue
        if path.name in assigned_files:
            continue
        role, spec = classify(path)
        if role == "skip":
            manifest["skipped"].append(path.name)
        elif role == "unassigned" or spec is None:
            manifest["unassigned"].append(path.name)
        else:
            parts = role.split(":")
            if parts[0] == "markets":
                manifest["markets"].setdefault(parts[1], {})[parts[2]] = spec
            else:
                manifest[parts[0]][parts[1]] = spec

    # xlsx screening reports need their sheet name
    for m in manifest["markets"].values():
        sr = m.get("screening_report")
        if sr and sr["format"] == "xlsx" and "sheet" not in sr:
            try:
                wb = openpyxl.load_workbook(ws / "raw" / sr["file"], read_only=True)
                sr["sheet"] = wb.sheetnames[0]
            except Exception:  # noqa: BLE001
                pass

    with open(mpath, "w") as f:
        yaml.safe_dump(manifest, f, sort_keys=False)
    return manifest
